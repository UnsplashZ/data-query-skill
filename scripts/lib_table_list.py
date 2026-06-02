#!/usr/bin/env python3
"""Shared table-list parsing helpers."""

from __future__ import annotations

import csv
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any


IDENTIFIER_RE = re.compile(r"^`?([^`.]+)`?(?:\.`?([^`.]+)`?)?$")


@dataclass(frozen=True)
class TableRequest:
    raw: str
    database: str
    table: str
    canonical: str
    source: str
    row_number: int
    duplicate_of: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def slug_identifier(value: str) -> str:
    text = str(value or "").strip().strip("`").strip()
    return re.sub(r"\s+", "", text)


def parse_table_name(raw: str, default_database: str | None = None) -> tuple[str, str, str]:
    text = str(raw or "").strip()
    if not text:
        raise ValueError("empty table name")
    text = text.replace('"', "`")
    match = IDENTIFIER_RE.match(text)
    if not match:
        parts = [slug_identifier(part) for part in text.replace("`", "").split(".") if slug_identifier(part)]
        if len(parts) == 1:
            database, table = default_database or "", parts[0]
        elif len(parts) == 2:
            database, table = parts
        else:
            raise ValueError(f"unsupported table name: {raw}")
    elif match.group(2):
        database, table = slug_identifier(match.group(1)), slug_identifier(match.group(2))
    else:
        database, table = default_database or "", slug_identifier(match.group(1))
    if not table:
        raise ValueError(f"missing table in table name: {raw}")
    canonical = f"{database}.{table}" if database else table
    return database, table, canonical


def _read_txt(path: Path) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    for lineno, raw in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        rows.append((lineno, line.split("#", 1)[0].strip()))
    return rows


def _read_csv(path: Path, table_column: str | None = None) -> list[tuple[int, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        has_header = csv.Sniffer().has_header(sample) if sample.strip() else False
        if table_column or has_header:
            reader = csv.DictReader(f)
            if table_column and reader.fieldnames and table_column not in reader.fieldnames:
                raise ValueError(f"table column not found in CSV: {table_column}")
            column = table_column or (reader.fieldnames[0] if reader.fieldnames else None)
            if not column:
                return []
            return [(i + 2, str(row.get(column) or "").strip()) for i, row in enumerate(reader) if str(row.get(column) or "").strip()]
        reader = csv.reader(f)
        return [(i, str(row[0]).strip()) for i, row in enumerate(reader, start=1) if row and str(row[0]).strip()]


def _read_xlsx(path: Path, sheet: str | None = None, table_column: str | None = None) -> list[tuple[int, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse xlsx table lists. Install with: python -m pip install -r requirements.txt") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        rows = list(rows_iter)
        if not rows:
            return []
        column_index = 0
        start_index = 0
        if table_column:
            headers = [str(value or "").strip() for value in rows[0]]
            if table_column not in headers:
                raise ValueError(f"table column not found in xlsx: {table_column}")
            column_index = headers.index(table_column)
            start_index = 1
        out: list[tuple[int, str]] = []
        for offset, row in enumerate(rows[start_index:], start=start_index + 1):
            if column_index < len(row):
                value = str(row[column_index] or "").strip()
                if value:
                    out.append((offset, value))
        return out
    finally:
        wb.close()


def parse_table_list(
    path: Path,
    *,
    default_database: str | None = None,
    table_column: str | None = None,
    sheet: str | None = None,
) -> dict[str, Any]:
    source = path.expanduser().resolve()
    suffix = source.suffix.lower()
    if suffix == ".txt":
        raw_rows = _read_txt(source)
    elif suffix == ".csv":
        raw_rows = _read_csv(source, table_column)
    elif suffix in {".xlsx", ".xlsm"}:
        raw_rows = _read_xlsx(source, sheet, table_column)
    else:
        raise ValueError(f"unsupported table-list suffix: {source.suffix}")

    requests: list[TableRequest] = []
    duplicates: list[dict[str, Any]] = []
    seen: dict[str, str] = {}
    for row_number, raw in raw_rows:
        database, table, canonical = parse_table_name(raw, default_database)
        duplicate_of = seen.get(canonical)
        request = TableRequest(
            raw=raw,
            database=database,
            table=table,
            canonical=canonical,
            source=str(source),
            row_number=row_number,
            duplicate_of=duplicate_of,
        )
        requests.append(request)
        if duplicate_of:
            duplicates.append(request.to_dict())
        else:
            seen[canonical] = raw

    return {
        "source": str(source),
        "default_database": default_database or "",
        "requested": [item.to_dict() for item in requests],
        "canonical": sorted(seen),
        "duplicate_aliases": duplicates,
    }
